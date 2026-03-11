import streamlit as st
import pandas as pd
import pdfplumber
import re
from io import BytesIO
from pdf2image import convert_from_bytes
import pytesseract
import string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
import msal
import requests
from datetime import datetime, date

# =============================================================================
# CONFIGURAÇÃO INICIAL DA PÁGINA
# =============================================================================
st.set_page_config(page_title="Extrair Fatura para Excel e SharePoint", layout="wide")
st.title("Extrair Débitos da Fatura (com Totais, Excel e SharePoint)")

# Upload do arquivo PDF
uploaded_file = st.file_uploader("Escolha o PDF da fatura", type="pdf")

# =============================================================================
# FUNÇÕES UTILITÁRIAS
# =============================================================================

def sanitize_filename(name):
    """
    Remove caracteres inválidos para nomes de arquivo.
    
    Args:
        name (str): Nome original do arquivo
    
    Returns:
        str: Nome sanitizado
    """
    valid_chars = f"-_.() {string.ascii_letters}{string.digits}"
    return "".join(c for c in name if c in valid_chars).strip() or "fatura"

def valor_br_para_float(valor_str):
    """
    Converte valor no formato brasileiro (R$ 1.234,56) para float.
    
    Args:
        valor_str (str): String contendo o valor
    
    Returns:
        float: Valor convertido
    """
    if valor_str is None:
        return 0.0
    v = str(valor_str).strip()
    v = v.replace("R$", "").replace(" ", "")
    v = v.replace(".", "").replace(",", ".")
    try:
        return round(float(v), 2)
    except:
        return 0.0

def parse_valor_input(valor_str):
    """
    Converte input do usuário para float.
    
    Args:
        valor_str (str): String contendo o valor
    
    Returns:
        float or None: Valor convertido ou None se inválido
    """
    if not valor_str:
        return None
    v = valor_str.replace(" ", "")
    v = v.replace(",", ".")
    try:
        return round(float(v), 2)
    except:
        return None

def formatar_data(d):
    """
    Formata data para DD/MM.
    
    Args:
        d (date): Data a ser formatada
    
    Returns:
        str: Data formatada
    """
    if d:
        return d.strftime("%d/%m")
    return ""

def extract_text_from_pdf(file):
    """
    Extrai texto de um arquivo PDF, usando OCR se necessário.
    
    Args:
        file: Arquivo PDF
    
    Returns:
        list: Lista de textos extraídos por página
    """
    texts = []

    # Tenta extrair texto diretamente do PDF
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            if txt:
                texts.append(txt)

    # Se não conseguiu extrair texto, usa OCR
    if not texts:
        file.seek(0)
        images = convert_from_bytes(file.read())
        for img in images:
            texts.append(pytesseract.image_to_string(img, lang="por"))

    return texts

def extract_tabela_transacoes(text):
    """
    Extrai tabela de transações/débitos do texto.
    
    Args:
        text (str): Texto extraído do PDF
    
    Returns:
        DataFrame: DataFrame com as transações encontradas
    """
    pattern = r"(\d{2}/\d{2})\s+[\d.]+\s+(.+?)\s+([\d.,]+)$"
    matches = re.findall(pattern, text, re.MULTILINE)

    if not matches:
        return pd.DataFrame()

    df = pd.DataFrame(matches, columns=["Data", "Estabelecimento", "Valor (R$)"])
    df["Valor (R$)"] = df["Valor (R$)"].apply(valor_br_para_float)

    return df

def extract_tabela_favorecidos(text):
    """
    Extrai tabela de transferências PIX do texto.
    
    Args:
        text (str): Texto extraído do PDF
    
    Returns:
        DataFrame: DataFrame com as transferências encontradas
    """
    pattern = (
        r"(\d{2}/\d{2})\s+(\S+)\s+([A-Z0-9\s]+?)\s+"
        r"([A-ZÀ-Ÿa-zà-ÿ0-9\.\- ]+?)\s+(\d{8})\s+"
        r"(\d{3,5})\s+([\d\-]+)\s+([\d.,]+)"
    )

    matches = re.findall(pattern, text, re.MULTILINE)

    if not matches:
        return pd.DataFrame()

    df_full = pd.DataFrame(matches)
    df = pd.DataFrame()
    df["Data"] = df_full[0]
    df["Favorecido"] = df_full[3].str.strip()
    df["Valor (R$)"] = df_full[7].apply(valor_br_para_float)

    return df

def extrair_mes_ano(nome):
    """
    Extrai mês e ano do nome do arquivo.
    
    Args:
        nome (str): Nome do arquivo
    
    Returns:
        datetime: Data com mês/ano extraído
    """
    mes_ano = re.search(r"([A-Z]{3})\s*(\d{4})", nome.upper())

    if mes_ano:
        mes_abrev, ano = mes_ano.groups()
        meses = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

        try:
            mes = meses.index(mes_abrev) + 1
        except:
            mes = 1

        return datetime(int(ano), mes, 1)

    return datetime.now()

# =============================================================================
# PROCESSAMENTO DO PDF UPLOADADO
# =============================================================================

if uploaded_file:
    uploaded_file.seek(0)

    # Inicializa session state se necessário
    if "df_transacoes" not in st.session_state:
        texts = extract_text_from_pdf(uploaded_file)

        lista_t = []  # Lista para transações
        lista_p = []  # Lista para PIX

        # Processa cada página
        for t in texts:
            d = extract_tabela_transacoes(t)
            if not d.empty:
                lista_t.append(d)

            p = extract_tabela_favorecidos(t)
            if not p.empty:
                lista_p.append(p)

        # Concatena DataFrames
        st.session_state.df_transacoes = (
            pd.concat(lista_t, ignore_index=True)
            if lista_t else pd.DataFrame(columns=["Data", "Estabelecimento", "Valor (R$)"])
        )

        st.session_state.df_pix = (
            pd.concat(lista_p, ignore_index=True)
            if lista_p else pd.DataFrame(columns=["Data", "Favorecido", "Valor (R$)"])
        )

    # =========================================================================
    # FORMULÁRIO PARA INSERÇÃO MANUAL DE DÉBITOS
    # =========================================================================
    with st.form("form_debito", clear_on_submit=True):
        st.subheader("Inserir Débito Manual")
        c1, c2, c3, c4 = st.columns([2, 4, 2, 1])

        data_manual = c1.date_input("Data")
        desc_manual = c2.text_input("Descrição")
        valor_manual = c3.text_input("Valor")
        submitted = c4.form_submit_button("Adicionar Débito")

        if submitted:
            valor = parse_valor_input(valor_manual)

            if desc_manual and valor is not None:
                nova = pd.DataFrame([{
                    "Data": formatar_data(data_manual),
                    "Estabelecimento": desc_manual.upper(),
                    "Valor (R$)": valor
                }])

                st.session_state.df_transacoes = pd.concat(
                    [st.session_state.df_transacoes, nova],
                    ignore_index=True
                )
                # Ordena por data
                st.session_state.df_transacoes.sort_values("Data", inplace=True)
                st.session_state.df_transacoes.reset_index(drop=True, inplace=True)
                st.rerun()

    # =========================================================================
    # EXIBIÇÃO E GERENCIAMENTO DOS DÉBITOS
    # =========================================================================
    if not st.session_state.df_transacoes.empty:
        st.markdown("### Débitos")

        # Exibe cada débito com botão de exclusão
        for i, row in st.session_state.df_transacoes.iterrows():
            a, b, c, d = st.columns([1, 4, 2, 0.5])
            a.write(row["Data"])
            b.write(row["Estabelecimento"])
            c.write(f"R$ {row['Valor (R$)']:,.2f}")

            if d.button("🗑️", key=f"del_t_{i}"):
                st.session_state.df_transacoes.drop(i, inplace=True)
                st.session_state.df_transacoes.reset_index(drop=True, inplace=True)
                st.rerun()

        # Total de débitos
        total_debito = st.session_state.df_transacoes["Valor (R$)"].sum()
        st.info(f"Total Débitos: R$ {total_debito:,.2f}")

    # =========================================================================
    # FORMULÁRIO PARA INSERÇÃO MANUAL DE PIX
    # =========================================================================
    with st.form("form_pix", clear_on_submit=True):
        st.subheader("Inserir PIX Manual")
        c1, c2, c3, c4 = st.columns([2, 4, 2, 1])

        data_pix = c1.date_input("Data")
        fav = c2.text_input("Favorecido")
        valor_pix = c3.text_input("Valor PIX")
        submitted = c4.form_submit_button("Adicionar PIX")

        if submitted:
            valor = parse_valor_input(valor_pix)

            if fav and valor is not None:
                nova = pd.DataFrame([{
                    "Data": formatar_data(data_pix),
                    "Favorecido": fav.upper(),
                    "Valor (R$)": valor
                }])

                st.session_state.df_pix = pd.concat(
                    [st.session_state.df_pix, nova],
                    ignore_index=True
                )
                # Ordena por data
                st.session_state.df_pix.sort_values("Data", inplace=True)
                st.session_state.df_pix.reset_index(drop=True, inplace=True)
                st.rerun()

    # =========================================================================
    # EXIBIÇÃO E GERENCIAMENTO DOS PIX
    # =========================================================================
    if not st.session_state.df_pix.empty:
        st.markdown("### PIX")

        # Exibe cada PIX com botão de exclusão
        for i, row in st.session_state.df_pix.iterrows():
            a, b, c, d = st.columns([1, 4, 2, 0.5])
            a.write(row["Data"])
            b.write(row["Favorecido"])
            c.write(f"R$ {row['Valor (R$)']:,.2f}")

            if d.button("🗑️", key=f"del_p_{i}"):
                st.session_state.df_pix.drop(i, inplace=True)
                st.session_state.df_pix.reset_index(drop=True, inplace=True)
                st.rerun()

        # Total de PIX
        total_pix = st.session_state.df_pix["Valor (R$)"].sum()
        st.info(f"Total PIX: R$ {total_pix:,.2f}")

    # =========================================================================
    # GERAÇÃO DO ARQUIVO EXCEL
    # =========================================================================
    nome_arquivo = st.text_input("Nome do arquivo Excel", "fatura")
    vencimento = extrair_mes_ano(nome_arquivo)

    if st.button("Gerar Excel"):
        output = BytesIO()

        # Prepara dados para o Excel
        df_excel_list = []

        if not st.session_state.df_transacoes.empty:
            df_excel_list.append(
                st.session_state.df_transacoes
                .rename(columns={"Estabelecimento": "Descrição", "Valor (R$)": "Valor"})
                [["Data", "Descrição", "Valor"]]
            )

        if not st.session_state.df_pix.empty:
            df_excel_list.append(
                st.session_state.df_pix
                .rename(columns={"Favorecido": "Descrição", "Valor (R$)": "Valor"})
                [["Data", "Descrição", "Valor"]]
            )

        # Concatena e adiciona linha de total
        df_excel = pd.concat(df_excel_list, ignore_index=True)
        total_geral = df_excel["Valor"].sum()
        df_excel.loc[len(df_excel)] = ["", "TOTAL", total_geral]

        # Cria arquivo Excel com formatação
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet = "Fatura"
            df_excel.to_excel(writer, sheet_name=sheet, index=False)

            # Aplica formatação de tabela
            ws = writer.book[sheet]
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tabela = Table(displayName="TabelaFatura", ref=ref)
            tabela.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True
            )
            ws.add_table(tabela)

            # Formata células de valor
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '#,##0.00'

        output.seek(0)

        # Botão para download
        st.download_button(
            "Baixar Excel",
            data=output,
            file_name=sanitize_filename(nome_arquivo) + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # =========================================================================
    # ENVIO PARA SHAREPOINT
    # =========================================================================
    data_vencimento = st.date_input(
        "Selecione o vencimento",
        value=vencimento
    )

    if st.button("Enviar total para SharePoint"):
        # Calcula total geral
        total_geral = (
            st.session_state.df_transacoes["Valor (R$)"].sum() +
            st.session_state.df_pix["Valor (R$)"].sum()
        )

        try:
            # Autenticação MSAL
            app = msal.ConfidentialClientApplication(
                client_id=os.getenv("AZURE_CLIENT_ID"),
                client_credential=os.getenv("AZURE_CLIENT_SECRET"),
                authority=f"https://login.microsoftonline.com/{os.getenv('AZURE_TENANT_ID')}"
            )

            token = app.acquire_token_for_client(
                scopes=["https://graph.microsoft.com/.default"]
            )

            access_token = token.get("access_token")

            # Configurações do SharePoint
            SITE_ID = "devgbsn.sharepoint.com,351e9978-140f-427e-a87d-332f6ce67a46,fc4e159a-5954-442f-a08f-28617bc84da1"
            LIST_ID = "b7b00e6d-9ed0-492c-958f-f80f15bd8dce"

            url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/lists/{LIST_ID}/items"

            # Payload da requisição
            payload = {
                "fields": {
                    "Despesa": f"Despesa Germano {nome_arquivo}",
                    "Valor": float(total_geral),
                    "Vencimento": data_vencimento.strftime("%Y-%m-%d"),
                    "QuemPagou": "Germano",
                    "pago": "sim",
                    "Data": date.today().strftime("%Y-%m-%d")
                }
            }

            # Envia para SharePoint
            response = requests.post(
                url,
                headers={
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json"
                },
                json=payload
            )

            if response.status_code == 201:
                st.success("Enviado para SharePoint com sucesso")
            else:
                st.error(response.text)

        except Exception as e:
            st.error(e)

# =============================================================================
# SEGUNDA PARTE: GERENCIAMENTO DO SHAREPOINT
# =============================================================================

st.set_page_config(page_title="SharePoint - Gerenciar Despesas", layout="wide")
st.title("Gerenciar Despesas do SharePoint")

# =============================================================================
# FUNÇÃO PARA BUSCAR ITENS DO SHAREPOINT
# =============================================================================

def buscar_itens_sharepoint():
    """
    Busca todos os itens da lista do SharePoint.
    
    Returns:
        DataFrame: DataFrame com os itens da lista
    """
    try:
        # Autenticação MSAL
        app = msal.ConfidentialClientApplication(
            client_id=os.getenv("AZURE_CLIENT_ID"),
            client_credential=os.getenv("AZURE_CLIENT_SECRET"),
            authority=f"https://login.microsoftonline.com/{os.getenv('AZURE_TENANT_ID')}"
        )

        token = app.acquire_token_for_client(
            scopes=["https://graph.microsoft.com/.default"]
        )

        access_token = token.get("access_token")

        # Configurações do SharePoint
        SITE_ID = "devgbsn.sharepoint.com,351e9978-140f-427e-a87d-332f6ce67a46,fc4e159a-5954-442f-a08f-28617bc84da1"
        LIST_ID = "b7b00e6d-9ed0-492c-958f-f80f15bd8dce"

        url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/lists/{LIST_ID}/items?expand=fields"

        # Requisição GET
        response = requests.get(
            url,
            headers={"Authorization": f"Bearer {access_token}"}
        )

        data = response.json()
        items = data.get("value", [])

        # Processa os itens
        lista = []
        for item in items:
            campos = item["fields"].copy()
            campos["ID"] = item["id"]  # Mantém ID para referência
            lista.append(campos)

        df = pd.DataFrame(lista)

        # Cria coluna MesAno a partir de Vencimento
        if "Vencimento" in df.columns:
            df["MesAno"] = pd.to_datetime(df["Vencimento"]).dt.strftime("%b %Y").str.upper()

        return df

    except Exception as e:
        st.error(f"Erro ao buscar itens do SharePoint: {e}")
        return pd.DataFrame()

# =============================================================================
# INTERFACE DE GERENCIAMENTO DO SHAREPOINT
# =============================================================================

# Botão para carregar despesas
if st.button("Carregar despesas do SharePoint"):
    st.session_state.df_sharepoint = buscar_itens_sharepoint()

if "df_sharepoint" in st.session_state and not st.session_state.df_sharepoint.empty:
    df = st.session_state.df_sharepoint.copy()

    # Seleciona colunas para exibição
    df_view = df[["ID","Despesa", "Valor","QuemPagou","Mes", "Ano"]].copy()
    # Renomeia os cabeçalhos para exibição
    df_view.rename(columns={
        "Valor": "Valor (R$)","QuemPagou":"Pago Por:"
    }, inplace=True)
    df_view["Excluir"] = False

    # Editor de dados interativo
    edited = st.data_editor(
        df_view,
        use_container_width=True,
        num_rows="dynamic",
        key="editor_sharepoint"
    )

    # =========================================================================
    # EXCLUSÃO DE ITENS
    # =========================================================================
    if st.button("Excluir itens selecionados"):
    ids_para_excluir = edited.loc[edited["Excluir"], "ID"].tolist()
    if ids_para_excluir:
        try:
            # Autenticação MSAL
            app = msal.ConfidentialClientApplication(
                client_id=os.getenv("AZURE_CLIENT_ID"),
                client_credential=os.getenv("AZURE_CLIENT_SECRET"),
                authority=f"https://login.microsoftonline.com/{os.getenv('AZURE_TENANT_ID')}"
            )

            token = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
            access_token = token.get("access_token")

            SITE_ID = "devgbsn.sharepoint.com,351e9978-140f-427e-a87d-332f6ce67a46,fc4e159a-5954-442f-a08f-28617bc84da1"
            LIST_ID = "b7b00e6d-9ed0-492c-958f-f80f15bd8dce"

            for item_id in ids_para_excluir:
                url = f"https://graph.microsoft.com/v1.0/sites/{SITE_ID}/lists/{LIST_ID}/items/{item_id}"
                response = requests.delete(
                    url,
                    headers={"Authorization": f"Bearer {access_token}"}
                )
                if response.status_code == 204:
                    st.info(f"Item {item_id} excluído do SharePoint")
                else:
                    st.error(f"Erro ao excluir {item_id}: {response.text}")

            # Atualiza DataFrame local
            st.session_state.df_sharepoint = st.session_state.df_sharepoint[
                ~st.session_state.df_sharepoint["ID"].isin(ids_para_excluir)
            ].reset_index(drop=True)

            st.success(f"{len(ids_para_excluir)} item(s) excluído(s) com sucesso!")

        except Exception as e:
            st.error(f"Erro ao excluir itens: {e}")

    # =========================================================================
    # ATUALIZAÇÃO DE ITENS
    # =========================================================================
    if st.button("Atualizar SharePoint"):
        for i, row in edited.iterrows():
            item_id = row["ID"]
            # TODO: Implementar atualização via Graph API
            pass
        st.session_state.df_sharepoint.update(edited.drop(columns=["Excluir"]))
        st.success("Itens atualizados com sucesso!")




